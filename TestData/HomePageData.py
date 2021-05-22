import openpyxl


class HomePageData:

    test_HomePage_Data = [{"FirstName": "Abhishek", "Email": "abhishekdas92@hotmail.com", "gender": "Male"},{"FirstName": "Jhilik", "Email": "jhilik@hotmail.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        dictValue = {}
        dataBook = openpyxl.load_workbook("H:\\Automation_Testing-Python\\TestData.xlsx")
        sheet = dataBook.active
        for i in range(1, sheet.max_row + 1):
            if (sheet.cell(row=i, column=1)).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # For column Value
                    dictValue[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        print(dictValue)
        return [dictValue]