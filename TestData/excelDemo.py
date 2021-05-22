import openpyxl


dataBook = openpyxl.load_workbook("H:\\Automation_Testing-Python\\TestData.xlsx")
#Select Active Sheet then cell
sheet = dataBook.active
dictValue = {}
cell = sheet.cell(row=1,column=2)
print(cell.value)

sheet.cell(row=2,column=2).value = "AbhishekDas"
print(sheet.cell(row=2,column=2).value)

#Total no of rows/columns in sheet present

print(sheet.max_row)
print(sheet.max_column)

#print(sheet['A3'].value)

#For loop for print each value in excel
list_val = []
for i in range(1,sheet.max_row+1):
    if (sheet.cell(row=i, column=1)).value == "TestCase2":
        for j in range(2, sheet.max_column+1):#For column Value
            dictValue[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dictValue)
